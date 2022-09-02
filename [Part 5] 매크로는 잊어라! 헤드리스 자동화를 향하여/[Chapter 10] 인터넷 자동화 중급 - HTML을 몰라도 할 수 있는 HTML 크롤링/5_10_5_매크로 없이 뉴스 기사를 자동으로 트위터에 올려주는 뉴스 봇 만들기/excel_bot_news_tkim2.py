from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup as BS
import pyautogui as pag
import pyperclip
import time
from pathlib import Path
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment

class NewsBot():
    def __init__(self):
        self.options = Options()
        self.options.add_argument('--window-size=1600,900')
        self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), chrome_options=self.options)
        self.news_list = []
        self.news_query = f'https://www.google.com/search?tbm=nws&q='
    
    def news_search(self, keyword, number):
        self.news_list = []
        self.driver.get(self.news_query + keyword)
        page = self.driver.page_source
        bs_obj = BS(page, 'html.parser')
        for i in range(int(number)):
            news_obj = bs_obj.find_all('div', {'class':'xuvV6b BGxR7d'})
            for j, news in enumerate(news_obj):
                news_company = news_obj[j].find('div', {'class':'CEMjEf NUnG9d'}).text
                news_title = news_obj[j].find('div', {'class':'mCBkyc y355M ynAwRc MBeuO nDgy9d'}).text
                news_summary = news_obj[j].find('div', {'class':'GI74Re nDgy9d'}).text.replace('\n', '').replace('\'', '"')
                news_time = news_obj[j].find('div', {'class':'OSrXXb ZE0LJd'}).text
                # 링크를 추출하고 싶은 경우 .attrs['href'] 등의 속성 활용 가능
                news_link = news_obj[j].find('a', {'class':'WlydOe'}).attrs['href']
                self.news_list.append([news_company, news_title, news_summary, news_time, news_link])
            # (By.TAG_NAME, 'html')을 활용해야 아래 쪽을 내려갈 수 있음
            html_element = self.driver.find_element(By.TAG_NAME, 'html')
            html_element.send_keys(Keys.END)
            time.sleep(1)
            next_link = self.driver.find_element(By.LINK_TEXT, '다음')
            next_link.click()
            time.sleep(1)

    def to_excel(self, filename):
        wb = openpyxl.Workbook()
        wb.create_sheet('Sheet')
        sheet = wb.active
        for index, content in enumerate(['언론사', '제목', '요약', '날짜', '링크']):
            sheet.cell(1, index+1).value = content
        for i, news in enumerate(self.news_list):
            for j, sub_news in enumerate(news):
                sheet.cell(i+2, j+1).value = sub_news
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 50
        sheet.column_dimensions['C'].width = 80
        bold_font = Font(bold=True)
        alignment_cell = Alignment(horizontal='center')
        for x in range(1, sheet.max_column + 1):
            sheet.cell(1, x).font = bold_font
            sheet.cell(1, x).alignment = alignment_cell
        wb.save(Path.cwd() / f'{filename}.xlsx')