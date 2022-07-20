from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup as BS
import pyautogui as pag
import time
import pyperclip
from pathlib import Path
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment

# jupyter notebook에서 일단 먼저 실험하는 중 - BeautifulSoup와 Selenium의 연계로
# BeautifulSoup와 Selenium의 연계로 나만의 코드로 진행

class NewsBot():
    def __init__(self):
        self.options = Options()
        self.options.add_argument('--window-size=1600,900')
        self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), chrome_options=self.options)
        # 'self.'의 여부 -> 다른 메서드로 이어질 것인가? 그냥 news_list로 하니까 안되고 self.news_list로 하니까 구동
        self.news_list = []
        self.news_query = 'https://www.google.com/search?tbm=nws&q='

    def kill(self):
        self.driver.quit()

# news_search 메서드로 페이지 이동까지 다 구현 -> 하나의 메서드로 하나의 큰 행동을 구현하는게 좋을듯
    def news_search(self, keyword, number):
        # self.news_list를 위에서 가져오는 형태?
        self.news_list = []
        self.driver.get(self.news_query + keyword)
        # page를 가져와서 BeautifulSoup 활용
        page = self.driver.page_source
        self.bs_obj = BS(page, 'html.parser')
        # 페이지를 몇 번 반복하면서 수집할 지
        for i in range(int(number)):
            self.news_obj = self.bs_obj.find_all('div', {'class':'xuvV6b BGxR7d'})
            # 해당 self.news_obj에서 언론사, 제목, 요약, n분 전 등을 각각 수집하여 리스트화
            for j, news in enumerate(self.news_obj):
                news_company = self.news_obj[j].find('div', {'class':'CEMjEf NUnG9d'}).text
                news_title = self.news_obj[j].find('div', {'class':'mCBkyc y355M ynAwRc MBeuO nDgy9d'}).text.replace('\'', '"')
                news_summary = self.news_obj[j].find('div', {'class':'GI74Re nDgy9d'}).text.replace('\n', '').replace('\'', '"')
                news_period = self.news_obj[j].find('div', {'class':'OSrXXb ZE0LJd'}).text
                self.news_list.append([news_company, news_title, news_summary, news_period])
            # selenium 활용하여 스크롤을 아래로 이동
            html_element = self.driver.find_element(By.TAG_NAME, 'html')
            html_element.send_keys(Keys.END)
            time.sleep(1)
            # selenium 활용하여 다음 페이지로 이동
            next_link = self.driver.find_element(By.LINK_TEXT, '다음')
            next_link.click()

    def to_excel(self, filename):
        wb = openpyxl.Workbook()
        wb.create_sheet('Sheet')
        sheet = wb.active
        for index, content in enumerate(['언론사', '제목', '요약']):
            sheet.cell(1, index+1).value = content
        for i, news in enumerate(self.news_list):
            for j, sub_news in enumerate(news):
                sheet.cell(i+2, j+1).value = sub_news
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 50
        sheet.column_dimensions['C'].width = 80
        bold_font = Font(bold=True)
        alignment_cell = Alignment(horizontal='center')
        for x in range(1, sheet.max_column + 1):
            sheet.cell(1, x).font = bold_font
            sheet.cell(1, x).alignment = alignment_cell
        wb.save(Path.cwd() / f'{filename}.xlsx')