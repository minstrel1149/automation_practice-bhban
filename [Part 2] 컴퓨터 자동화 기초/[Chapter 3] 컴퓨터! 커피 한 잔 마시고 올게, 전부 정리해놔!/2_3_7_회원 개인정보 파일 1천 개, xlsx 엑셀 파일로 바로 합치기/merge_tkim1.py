#-*-coding:utf-8

from pathlib import Path
import os
import shutil
# 학습한 적 있는 openpyxl로 진행 도전 -> 엑셀 관련은 무조건 openpyxl로
import openpyxl

# 폴더를 복사하는 것부티 진행 -> shutil모듈 활용(copytree)
p1 = Path.cwd() / '[Part 2] 컴퓨터 자동화 기초' / '[Chapter 3] 컴퓨터! 커피 한 잔 마시고 올게, 전부 정리해놔!' / '2_3_4_개인정보 파일 1천 개, CSV 파일로 예쁘게 합치기' / 'personal_info_tkim1'
p2 = Path.cwd() / '[Part 2] 컴퓨터 자동화 기초' / '[Chapter 3] 컴퓨터! 커피 한 잔 마시고 올게, 전부 정리해놔!' / '2_3_7_회원 개인정보 파일 1천 개, xlsx 엑셀 파일로 바로 합치기' / 'personal_info_tkim1'
p3 = Path.cwd() / '[Part 2] 컴퓨터 자동화 기초' / '[Chapter 3] 컴퓨터! 커피 한 잔 마시고 올게, 전부 정리해놔!' / '2_3_7_회원 개인정보 파일 1천 개, xlsx 엑셀 파일로 바로 합치기'
try:
    shutil.copytree(p1, p2)
except:
    print('이미 존재하는 폴더입니다.')

outfile_name = 'merged_ID_tkim1.xlsx'

files = os.listdir(p2)

# 엑셀 워크북 생성할 때 괄호 안에 경로를 지정해주면 안됨! -> 각 셀 값들 변환이 불가능!
new_file = openpyxl.Workbook()
try:
    new_file.create_sheet(index=0, title='Sheet')
except:
    print('이미 존재하는 시트입니다.')

new_sheet = new_file.active

# # headers 빈 리스트 생성
headers = []
# filename in files 대신 index를 활용하는 방식(excel 각 행마다 contents를 삽입하기 위해)
for i in range(len(files)):
    if '.txt' not in files[i]:
        continue
    read_file = open(p2 / files[i], mode='r')
    # contents를 담을 빈 리스트 생성
    contents = []
    # readlines 메서드로 한 줄씩 반복
    for line in read_file.readlines():
        splits = line.split(':')
        contents.append(splits[-1].strip())
        # headers는 단 한번만 생성(빈 리스트에 추가하고 나면 contents와 header의 길이가 같아짐)
        if len(contents) > len(headers):
            headers.append(splits[0].strip())
    # 리스트에 추가된 contents 자료를 excel에 업데이트
    for j in range(len(contents)):
        new_sheet.cell(i+2, j+1).value = contents[j]

for i in range(len(headers)):
    new_sheet.cell(1, i+1).value = headers[i]


new_file.save(p3 / outfile_name)

print('Process Done!')