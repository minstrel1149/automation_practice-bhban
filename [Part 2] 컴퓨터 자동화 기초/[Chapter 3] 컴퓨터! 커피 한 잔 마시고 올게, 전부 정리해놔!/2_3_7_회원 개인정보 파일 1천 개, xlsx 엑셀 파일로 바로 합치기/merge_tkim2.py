from pathlib import Path
import os
import shutil
import openpyxl

p1 = Path.home() / 'bhban_rpa-master' / '[Part 2] 컴퓨터 자동화 기초' / '[Chapter 3] 컴퓨터! 커피 한 잔 마시고 올게, 전부 정리해놔!' / '2_3_4_개인정보 파일 1천 개, CSV 파일로 예쁘게 합치기' / 'personal_info_tkim2'
p2 = Path.home() / 'bhban_rpa-master' / '[Part 2] 컴퓨터 자동화 기초' / '[Chapter 3] 컴퓨터! 커피 한 잔 마시고 올게, 전부 정리해놔!' / '2_3_7_회원 개인정보 파일 1천 개, xlsx 엑셀 파일로 바로 합치기' / 'personal_info_tkim2'
p3 = Path.home() / 'bhban_rpa-master' / '[Part 2] 컴퓨터 자동화 기초' / '[Chapter 3] 컴퓨터! 커피 한 잔 마시고 올게, 전부 정리해놔!' / '2_3_7_회원 개인정보 파일 1천 개, xlsx 엑셀 파일로 바로 합치기'
try:
    shutil.copytree(p1, p2)
except:
    print('이미 존재하는 폴더입니다.')

files = os.listdir(p2)
# os.listdir() 대신 Path객체의 glob() 메서드 및 list comprehension 활용 가능
# p2.glob('*.txt')는 Generator 객체(Lazy Iterator) 반환
files2 = [x.name for x in p2.glob('*.txt')]

wb = openpyxl.Workbook()
try:
    new_file.create_sheet(index=0, title='Sheet')
except:
    print('이미 존재하는 시트입니다.')

sheet = wb.active

out_file_has_header = False
headers = []
for i, filename in enumerate(files):
    if '.txt' not in filename:
        continue
    read_file = open(p2 / filename, mode='r')
    contents = []
    for line in read_file.read().splitlines():
        contents.append(line.split(':')[-1].strip())
        if len(contents) > len(headers):
            headers.append(line.split(':')[0].strip())
    for j, content in enumerate(contents):
        sheet.cell(i+2, j+1).value = content

for k, header in enumerate(headers):
    sheet.cell(1, k+1).value = header

wb.save(p3 / 'merged_ID_tkim2.xlsx')

print('Process Done!')
