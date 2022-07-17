from pathlib import Path
import time
import random
import shutil
import os
# openpyxl모듈을 활용하여 생성 시도 -> 책과는 관계 없이 내가 생각하는 방향으로
import openpyxl

# 폴더를 복사하는 것부티 진행 -> shutil모듈 활용(copytree)
p1 = Path.cwd() / '[Part 2] 컴퓨터 자동화 기초' / '[Chapter 4] 컴퓨터! 엑셀 정리좀 해 놔, 10초 안에!' / '2_4_4_예제용 엑셀파일 1천 개, 엔터키 한 번에 만들기' / 'personal_info_tkim1'
p2 = Path.cwd() / '[Part 2] 컴퓨터 자동화 기초' / '[Chapter 4] 컴퓨터! 엑셀 정리좀 해 놔, 10초 안에!' / '2_4_5_엑셀파일 1천 개, 순식간에 하나로 합치기' / 'personal_info_tkim1'
p3 = Path.cwd() / '[Part 2] 컴퓨터 자동화 기초' / '[Chapter 4] 컴퓨터! 엑셀 정리좀 해 놔, 10초 안에!' / '2_4_5_엑셀파일 1천 개, 순식간에 하나로 합치기'
try:
    shutil.copytree(p1, p2)
except:
    print('이미 존재하는 폴더입니다.')

files = os.listdir(p2)
outfile_name = 'merged_ID_tkim1.xlsx'

write_file = openpyxl.Workbook()
write_file.create_sheet
write_sheet = write_file.active

header = False
for i, filename in enumerate(files):
    read_files = openpyxl.load_workbook(p2 / filename)
    read_sheets = read_files.active
    if header == False:
        for header_i in range(read_sheets.max_column):
            # openpyxl도 list 배열 형태로 업데이트가 가능하니, 다음 번에는 그걸로 시도
            write_sheet.cell(1, header_i+1).value = read_sheets.cell(1, header_i+1).value
        header = True
    for j in range(read_sheets.max_column):
        write_sheet.cell(i+2, j+1).value = read_sheets.cell(2, j+1).value

write_file.save(p3 / outfile_name)

print('Process Done!')
