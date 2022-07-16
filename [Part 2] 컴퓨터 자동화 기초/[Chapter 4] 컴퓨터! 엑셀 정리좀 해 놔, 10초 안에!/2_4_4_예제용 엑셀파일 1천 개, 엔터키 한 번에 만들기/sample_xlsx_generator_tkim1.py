# 여기 있던 utf-8 관련 내용을 지우니까 syntax error가 없어짐.

from pathlib import Path
import time
import random
import os
# 책 모듈 대신 openpyxl을 활용하여 생성 시도
import openpyxl

num_samples = 1000
alphabet_samples = 'abcdefghijklmnopqrstuvwxyz1234567890'

def random_string(length):
    result = ''
    for i in range(length):
        result += random.choice(alphabet_samples)
    return result

first_name_samples = '김이박최정강조윤장임'
middle_name_samples = '민서예지도하주윤채현지'
last_name_samples = '준윤우원호후서연아은진'

def random_name():
    result = ''
    result += random.choice(first_name_samples)
    result += random.choice(middle_name_samples)
    result += random.choice(last_name_samples)
    return result

try:
    (Path.cwd() / '[Part 2] 컴퓨터 자동화 기초' / '[Chapter 4] 컴퓨터! 엑셀 정리좀 해 놔, 10초 안에!' / '2_4_4_예제용 엑셀파일 1천 개, 엔터키 한 번에 만들기' / 'personal_info_tkim1').mkdir()
except:
    print('이미 존재하는 폴더입니다.')

p1 = Path.cwd() / '[Part 2] 컴퓨터 자동화 기초' / '[Chapter 4] 컴퓨터! 엑셀 정리좀 해 놔, 10초 안에!' / '2_4_4_예제용 엑셀파일 1천 개, 엔터키 한 번에 만들기' / 'personal_info_tkim1'

for i in range(num_samples):
    # 괄호가 꼭 있어야!
    name = random_name()
    filename = p1 / f'{i+1}_{name}.xlsx'
    header = ['이름', '나이', '메일', '전화', '성별']
    age = random.randint(20, 39)
    email = f'{random_string(8)}@taehwakim.co.kr'
    telephone = f'010-{random.randint(1000, 9999)}-{random.randint(1000, 9999)}'
    sex = random.choice(['남성', '여성'])
    # 이름부터 성별까지를 contents라는 이름의 list로 처리
    contents = [name, age, email, telephone, sex]
    # openpyxl로 새로운 excel파일 생성
    new_file = openpyxl.Workbook()
    new_file.create_sheet()
    new_sheet = new_file.active
    # 1행에 header 삽입
    for i in range(len(header)):
        new_sheet.cell(1, i+1).value = header[i]
    # 2행에 contents 삽입
    for i in range(len(contents)):
        new_sheet.cell(2, i+1).value = contents[i]
    new_file.save(filename)

print('Process Done!')
