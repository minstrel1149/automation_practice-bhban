from pathlib import Path
import os
import random
import openpyxl
import string

num_samples = 1000
alphabet_samples = string.ascii_lowercase + string.digits

def random_string(length):
    result = ''
    for i in range(length):
        result += random.choice(alphabet_samples)
    return result

first_name_samples = '김이박최정강조윤장임'
middle_name_samples = '민서예지도하주윤채현지'
last_name_samples = '준윤우원호후서연아은진'

def random_name():
    result = ''
    result += random.choice(first_name_samples)
    result += random.choice(middle_name_samples)
    result += random.choice(last_name_samples)
    return result

try:
    (Path.home() / 'bhban_rpa-master' / '[Part 2] 컴퓨터 자동화 기초' / '[Chapter 4] 컴퓨터! 엑셀 정리좀 해 놔, 10초 안에!' / '2_4_4_예제용 엑셀파일 1천 개, 엔터키 한 번에 만들기' / 'personal_info_tkim2').mkdir()
except:
    print('이미 존재하는 폴더입니다.')

p = Path.home() / 'bhban_rpa-master' / '[Part 2] 컴퓨터 자동화 기초' / '[Chapter 4] 컴퓨터! 엑셀 정리좀 해 놔, 10초 안에!' / '2_4_4_예제용 엑셀파일 1천 개, 엔터키 한 번에 만들기' / 'personal_info_tkim2'

for i in range(num_samples):
    header = ['이름', '나이', '이메일', '전화번호', '성별']
    name = random_name()
    age = random.randint(20, 39)
    email = f'{random_string(random.randint(5, 9))}@taehwakim.co.kr'
    phone = f'010-{random.randint(1111, 9999)}-{random.randint(1111, 9999)}'
    sex = random.choice(['남자', '여자'])
    contents = [name, age, email, phone, sex]
    wb = openpyxl.Workbook()
    sheet = wb.active
    for j, (head, content) in enumerate(zip(header, contents)):
        sheet.cell(1, j+1).value = head
        sheet.cell(2, j+1).value = content
    wb.save(p / f'{i+1}_{name}.xlsx')

print('Process Done!')