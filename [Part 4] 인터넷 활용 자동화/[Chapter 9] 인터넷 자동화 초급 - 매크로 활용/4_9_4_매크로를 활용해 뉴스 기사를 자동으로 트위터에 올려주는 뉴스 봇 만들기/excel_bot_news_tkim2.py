from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
# webdriver_manager 인스톨 필요
from webdriver_manager.chrome import ChromeDriverManager
import pyautogui as pag
import time
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import pyperclip
from pathlib import Path

class NewsBot:
    def __init__(self):
        self.options = Options()
        self.options.add_argument('--window-size=1600,900')
        # webdriver.Chrome()에서 Chrome의 C는 대문자여야
        self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), chrome_options=self.options)
        self.news_query = f'https://www.google.com/search?tbm=nws&q='
        self.news_list = []
        self.news_text = ''
    
    def search(self, keyword):
        self.driver.get(f'{self.news_query}{keyword}')
        time.sleep(1)
    
    def refresh(self):
        pag.press('f5')
    
    def copy_all(self):
        pag.hotkey('ctrl', 'a')
        time.sleep(1)
        pag.hotkey('ctrl', 'c')
        time.sleep(1)
        pag.hotkey('ctrl', 'c')
    
    def scrap_news(self):
        self.copy_all()
        self.news_list = []
        self.news_text = pyperclip.paste()
        # list로 변환할 때 따옴표 때문에 걸리는게 있어서 모두 큰 따옴표로 전환
        split_text = self.news_text.replace('\'', '"').splitlines()
        for i, line in enumerate(split_text):
            if len(line.strip()) < 3:
                continue
            elif line.strip()[-4:] in '달 전 주 전 일 전 시간 전 분 전 초 전':
                # 간편 조건문을 통해 '이자 뉴스'가 있는 경우를 처리
                new_news_list = split_text[i-3:i] if split_text[i-3] != '' else split_text[i-2:i]
                self.news_list.append(new_news_list)
    
    def news_crawler(self, keyword):
        self.search(keyword)
        self.scrap_news()
    
    def to_excel(self, filename):
        wb = openpyxl.Workbook()
        wb.create_sheet('Sheet')
        sheet = wb.active
        for index, content in enumerate(['언론사', '제목', '요약']):
            sheet.cell(1, index+1).value = content
        for i, news in enumerate(self.news_list):
            for j, part in enumerate(news):
                sheet.cell(i+2, j+1).value = part
        # .width 빼먹지말고 써야! 아니면 LookupError: unknown encoding: us-ascii 오류 생김
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 50
        sheet.column_dimensions['C'].width = 80
        bold_font = Font(bold=True)
        alignment_cell = Alignment(horizontal='center')
        for x in range(1, sheet.max_column + 1):
            sheet.cell(1, x).font = bold_font
            sheet.cell(1, x).alignment = alignment_cell
        wb.save(Path.cwd() / f'{filename}.xlsx')

