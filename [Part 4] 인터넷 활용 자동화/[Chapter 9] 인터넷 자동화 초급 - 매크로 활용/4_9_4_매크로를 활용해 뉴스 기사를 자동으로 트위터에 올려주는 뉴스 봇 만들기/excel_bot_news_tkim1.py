from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
# webdriver_manager 인스톨 필요
from webdriver_manager.chrome import ChromeDriverManager
import pyautogui as pag
import time
import pyperclip
from pathlib import Path
import openpyxl
# openpyxl의 하위 기능 Font, Alignment 활용
from openpyxl.styles import Font
from openpyxl.styles import Alignment

class NewsBot:
    def __init__(self):
        # news_query 에서 검색어를 제외한 나머지 링크 작성
        self.news_query = 'https://www.google.com/search?tbm=nws&q='
        self.options = Options()
        self.options.add_argument('--window-size=1600,900')
        self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), chrome_options=self.options)
        self.news_list = []
        self.news_text = ''
    def quit_bot(self):
        self.driver.quit()
    def search(self, keyword):
        # 문자열을 합쳐줄 수 있는 '+'로 처리해야함을 잊지 말기
        self.driver.get(self.news_query + keyword)
        time.sleep(2)
    def refresh(self):
        # pyautogui 활용
        pag.press('f5')
    def copy_all(self):
        pag.hotkey('ctrl', 'a')
        time.sleep(1)
        pag.hotkey('ctrl', 'c')
        time.sleep(1)
        pag.hotkey('ctrl', 'c')
    def scrap_news(self):
        # copy_all 메서드를 활용해 먼저 복사 진행
        self.copy_all()
        self.news_list = []
        self.news_text = pyperclip.paste()
        # list로 변환할 때 따옴표 때문에 걸리는게 있어서 모두 큰 따옴표로 전환
        split_text = self.news_text.replace('\'', '"').splitlines()
        for i, line in enumerate(split_text):
            # 길이가 3 이하인 경우 빈 내용일 가능성이 높음을 고려
            if len(line.strip()) < 3:
                continue
            elif line.strip()[-3:] in '달 전 주 전 일 전 시간 전 분 전 초 전':
                # 엑셀에 저장할 것이기 때문에 join으로 합치지 말고 list 형태로 진행해야하는데..
                new_news_str = '\n'.join(split_text[i-3:i])
                # 그냥 splitlines()를 한번 더 써서 list로 전환
                new_news_list = new_news_str.splitlines()
                self.news_list.append(new_news_list)
    def news_crawler(self, keyword):
        self.search(keyword)
        self.scrap_news()
    def to_excel(self, filename):
        wb = openpyxl.Workbook()
        wb.create_sheet('Sheet')
        sheet = wb.active
        # 엑셀의 헤드 기재
        for index, content in enumerate(['언론사', '제목', '요약']):
            sheet.cell(1, index+1).value = content
        # 그냥 일단 중첩 for문을 통해 기재 -> 추후 array형태 고려 필요
        for i, news in enumerate(self.news_list):
            for j, sub_news in enumerate(news):
                # i+1 기재해줘야함을 헷갈리지 말고 처리해야
                sheet.cell(i+2, j+1).value = sub_news
        # column_dimensions 활용해 너비 조정
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 50
        sheet.column_dimensions['C'].width = 80
        # 헤드행 폰트를 bold=True로 지정
        bold_font = Font(bold=True)
        # 헤드행을 중앙정렬로 진행
        alignment_cell = Alignment(horizontal='center')
        for x in range(1, sheet.max_column + 1):
            sheet.cell(1, x).font = bold_font
            sheet.cell(1, x).alignment = alignment_cell
        wb.save(Path.cwd() / f'{filename}.xlsx')
